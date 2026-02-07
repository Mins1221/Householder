# -*- coding: utf-8 -*- 

###########################################################################
## Modern Smart Household Account Book
## 모던 스마트 가계부 v6.0 - 안정화 버전
###########################################################################

import wx
import wx.xrc
import wx.adv
import re
import csv
from datetime import datetime
from collections import defaultdict

# 모듈 import (실제 환경에 맞게 수정)
try:
    from main import HL_CRUD
    from main.barChart import Barchart
except ImportError:
    # 개발 환경용 더미 클래스
    class HL_CRUD:
        @staticmethod
        def selectMonthList():
            return ['2025-01', '2025-02', '2025-03']
        
        @staticmethod
        def selectAll():
            return []
        
        @staticmethod
        def selectMonthlySum(month):
            return [('', month, '합계', '', '0', '0', '')]
        
        @staticmethod
        def insert(data):
            pass
        
        @staticmethod
        def update(data):
            pass
        
        @staticmethod
        def delete(key):
            pass
    
    class Barchart(wx.Panel):
        def __init__(self, parent):
            super().__init__(parent)
            self.SetBackgroundColour(wx.WHITE)
            
        def SetData(self, data):
            pass


###########################################################################
## 색상 테마 설정 - 기본 안정 버전
###########################################################################
class ColorTheme:
    # 배경색
    BG_PRIMARY = wx.Colour(245, 245, 250)
    BG_SECONDARY = wx.WHITE
    
    # 패널 & 카드
    CARD_BG = wx.WHITE
    PANEL_BG = wx.Colour(250, 250, 252)
    
    # 텍스트
    TEXT_PRIMARY = wx.Colour(50, 50, 50)
    TEXT_SECONDARY = wx.Colour(100, 100, 100)
    
    # 액센트 컬러
    ACCENT = wx.Colour(0, 122, 255)
    ACCENT_LIGHT = wx.Colour(100, 160, 255)
    
    # 상태 컬러
    SUCCESS = wx.Colour(52, 199, 89)
    WARNING = wx.Colour(255, 149, 0)
    DANGER = wx.Colour(255, 59, 48)
    
    # 수입/지출
    INCOME_COLOR = wx.Colour(52, 199, 89)
    EXPENSE_COLOR = wx.Colour(255, 59, 48)
    
    # Border
    BORDER = wx.Colour(220, 220, 220)


###########################################################################
## 카드 패널
###########################################################################
class CardPanel(wx.Panel):
    def __init__(self, parent, title=""):
        super().__init__(parent)
        self.SetBackgroundColour(ColorTheme.CARD_BG)
        
        self.main_sizer = wx.BoxSizer(wx.VERTICAL)
        
        if title:
            title_text = wx.StaticText(self, label=title)
            font = title_text.GetFont()
            font.SetPointSize(14)
            font.SetWeight(wx.FONTWEIGHT_BOLD)
            title_text.SetFont(font)
            title_text.SetForegroundColour(ColorTheme.TEXT_PRIMARY)
            
            self.main_sizer.Add(title_text, 0, wx.ALL, 15)
            
            # 구분선
            line = wx.Panel(self, size=(-1, 1))
            line.SetBackgroundColour(ColorTheme.BORDER)
            self.main_sizer.Add(line, 0, wx.EXPAND | wx.LEFT | wx.RIGHT, 15)
        
        self.SetSizer(self.main_sizer)
    
    def AddContent(self, content):
        self.main_sizer.Add(content, 1, wx.EXPAND | wx.ALL, 15)


###########################################################################
## 검색 다이얼로그
###########################################################################
class SearchDialog(wx.Dialog):
    def __init__(self, parent):
        super().__init__(parent, title="고급 검색", size=(500, 500))
        self.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        
        panel = wx.Panel(self)
        panel.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 날짜 범위
        date_card = CardPanel(panel, "날짜 범위")
        date_content = wx.Panel(date_card)
        date_content.SetBackgroundColour(ColorTheme.CARD_BG)
        date_sizer = wx.FlexGridSizer(2, 2, 10, 10)
        
        date_sizer.Add(wx.StaticText(date_content, label="시작일:"), 0, wx.ALIGN_CENTER_VERTICAL)
        self.start_date = wx.adv.DatePickerCtrl(date_content, style=wx.adv.DP_DROPDOWN | wx.adv.DP_SHOWCENTURY)
        date_sizer.Add(self.start_date, 1, wx.EXPAND)
        
        date_sizer.Add(wx.StaticText(date_content, label="종료일:"), 0, wx.ALIGN_CENTER_VERTICAL)
        self.end_date = wx.adv.DatePickerCtrl(date_content, style=wx.adv.DP_DROPDOWN | wx.adv.DP_SHOWCENTURY)
        self.end_date.SetValue(wx.DateTime.Today())
        date_sizer.Add(self.end_date, 1, wx.EXPAND)
        
        date_content.SetSizer(date_sizer)
        date_card.AddContent(date_content)
        sizer.Add(date_card, 0, wx.EXPAND | wx.ALL, 10)
        
        # 구분
        type_card = CardPanel(panel, "거래 유형")
        type_content = wx.Panel(type_card)
        type_content.SetBackgroundColour(ColorTheme.CARD_BG)
        type_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        self.chk_income = wx.CheckBox(type_content, label="수입")
        self.chk_income.SetValue(True)
        self.chk_expense = wx.CheckBox(type_content, label="지출")
        self.chk_expense.SetValue(True)
        
        type_sizer.Add(self.chk_income, 0, wx.ALL, 5)
        type_sizer.Add(self.chk_expense, 0, wx.ALL, 5)
        
        type_content.SetSizer(type_sizer)
        type_card.AddContent(type_content)
        sizer.Add(type_card, 0, wx.EXPAND | wx.ALL, 10)
        
        # 금액 범위
        amount_card = CardPanel(panel, "금액 범위")
        amount_content = wx.Panel(amount_card)
        amount_content.SetBackgroundColour(ColorTheme.CARD_BG)
        amount_sizer = wx.FlexGridSizer(2, 2, 10, 10)
        
        amount_sizer.Add(wx.StaticText(amount_content, label="최소 금액:"), 0, wx.ALIGN_CENTER_VERTICAL)
        self.min_amount = wx.TextCtrl(amount_content)
        amount_sizer.Add(self.min_amount, 1, wx.EXPAND)
        
        amount_sizer.Add(wx.StaticText(amount_content, label="최대 금액:"), 0, wx.ALIGN_CENTER_VERTICAL)
        self.max_amount = wx.TextCtrl(amount_content)
        amount_sizer.Add(self.max_amount, 1, wx.EXPAND)
        
        amount_content.SetSizer(amount_sizer)
        amount_card.AddContent(amount_content)
        sizer.Add(amount_card, 0, wx.EXPAND | wx.ALL, 10)
        
        # 키워드
        keyword_card = CardPanel(panel, "키워드 검색")
        keyword_content = wx.Panel(keyword_card)
        keyword_content.SetBackgroundColour(ColorTheme.CARD_BG)
        keyword_sizer = wx.BoxSizer(wx.VERTICAL)
        
        self.keyword = wx.TextCtrl(keyword_content)
        keyword_sizer.Add(self.keyword, 0, wx.EXPAND)
        
        keyword_content.SetSizer(keyword_sizer)
        keyword_card.AddContent(keyword_content)
        sizer.Add(keyword_card, 0, wx.EXPAND | wx.ALL, 10)
        
        # 버튼
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        btn_sizer.AddStretchSpacer()
        
        btn_cancel = wx.Button(panel, wx.ID_CANCEL, "취소", size=(100, 36))
        btn_ok = wx.Button(panel, wx.ID_OK, "검색", size=(100, 36))
        btn_ok.SetBackgroundColour(ColorTheme.ACCENT)
        btn_ok.SetForegroundColour(wx.WHITE)
        
        btn_sizer.Add(btn_cancel, 0, wx.ALL, 5)
        btn_sizer.Add(btn_ok, 0, wx.ALL, 5)
        
        sizer.Add(btn_sizer, 0, wx.EXPAND | wx.ALL, 10)
        
        panel.SetSizer(sizer)
        
        # 날짜 초기화
        start = wx.DateTime.Today()
        start.SetDay(1)
        self.start_date.SetValue(start)
    
    def GetSearchCriteria(self):
        start = self.start_date.GetValue()
        end = self.end_date.GetValue()
        
        return {
            'start_date': start.FormatISODate(),
            'end_date': end.FormatISODate(),
            'include_income': self.chk_income.GetValue(),
            'include_expense': self.chk_expense.GetValue(),
            'min_amount': self.min_amount.GetValue(),
            'max_amount': self.max_amount.GetValue(),
            'keyword': self.keyword.GetValue()
        }


###########################################################################
## 통계 다이얼로그
###########################################################################
class StatisticsDialog(wx.Dialog):
    def __init__(self, parent, data):
        super().__init__(parent, title="통계 분석", size=(700, 600))
        self.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        
        panel = wx.Panel(self)
        panel.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 월별 통계
        monthly_data = defaultdict(lambda: {'income': 0, 'expense': 0})
        
        for row in data:
            month = row[1][:7]  # YYYY-MM
            if row[2] == '수입':
                monthly_data[month]['income'] += float(row[4]) if row[4] else 0
            else:
                monthly_data[month]['expense'] += float(row[5]) if row[5] else 0
        
        # 리스트 컨트롤
        list_card = CardPanel(panel, "월별 수입/지출 통계")
        list_ctrl = wx.ListCtrl(list_card, style=wx.LC_REPORT | wx.BORDER_SIMPLE)
        list_ctrl.SetBackgroundColour(wx.WHITE)
        
        list_ctrl.InsertColumn(0, "월", width=120)
        list_ctrl.InsertColumn(1, "수입", width=150)
        list_ctrl.InsertColumn(2, "지출", width=150)
        list_ctrl.InsertColumn(3, "수지", width=150)
        
        for month in sorted(monthly_data.keys(), reverse=True):
            income = monthly_data[month]['income']
            expense = monthly_data[month]['expense']
            balance = income - expense
            
            idx = list_ctrl.InsertItem(list_ctrl.GetItemCount(), month)
            list_ctrl.SetItem(idx, 1, f"{income:,.0f}원")
            list_ctrl.SetItem(idx, 2, f"{expense:,.0f}원")
            
            balance_text = f"{balance:,.0f}원"
            list_ctrl.SetItem(idx, 3, balance_text)
            
            if balance >= 0:
                list_ctrl.SetItemTextColour(idx, ColorTheme.INCOME_COLOR)
            else:
                list_ctrl.SetItemTextColour(idx, ColorTheme.EXPENSE_COLOR)
        
        list_card.AddContent(list_ctrl)
        sizer.Add(list_card, 1, wx.EXPAND | wx.ALL, 10)
        
        # 닫기 버튼
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        btn_sizer.AddStretchSpacer()
        
        btn_close = wx.Button(panel, wx.ID_CLOSE, "닫기", size=(100, 36))
        btn_sizer.Add(btn_close, 0, wx.ALL, 10)
        
        sizer.Add(btn_sizer, 0, wx.EXPAND)
        
        panel.SetSizer(sizer)
        
        self.Bind(wx.EVT_BUTTON, self.OnClose, btn_close)
    
    def OnClose(self, event):
        self.Close()


###########################################################################
## 예산 다이얼로그
###########################################################################
class BudgetDialog(wx.Dialog):
    def __init__(self, parent):
        super().__init__(parent, title="예산 관리", size=(500, 400))
        self.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        
        panel = wx.Panel(self)
        panel.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 안내 메시지
        info_card = CardPanel(panel, "예산 설정")
        info_text = wx.StaticText(info_card, label="월별 예산을 설정하고 지출을 관리하세요.")
        info_text.SetForegroundColour(ColorTheme.TEXT_SECONDARY)
        info_card.AddContent(info_text)
        sizer.Add(info_card, 0, wx.EXPAND | wx.ALL, 10)
        
        # 예산 입력
        budget_card = CardPanel(panel, "월 예산")
        budget_content = wx.Panel(budget_card)
        budget_content.SetBackgroundColour(ColorTheme.CARD_BG)
        budget_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        self.budget_input = wx.TextCtrl(budget_content, size=(200, -1))
        budget_sizer.Add(self.budget_input, 1, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 5)
        budget_sizer.Add(wx.StaticText(budget_content, label="원"), 0, wx.ALIGN_CENTER_VERTICAL)
        
        budget_content.SetSizer(budget_sizer)
        budget_card.AddContent(budget_content)
        sizer.Add(budget_card, 0, wx.EXPAND | wx.ALL, 10)
        
        # 버튼
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        btn_sizer.AddStretchSpacer()
        
        btn_cancel = wx.Button(panel, wx.ID_CANCEL, "취소", size=(100, 36))
        btn_save = wx.Button(panel, wx.ID_OK, "저장", size=(100, 36))
        btn_save.SetBackgroundColour(ColorTheme.ACCENT)
        btn_save.SetForegroundColour(wx.WHITE)
        
        btn_sizer.Add(btn_cancel, 0, wx.ALL, 5)
        btn_sizer.Add(btn_save, 0, wx.ALL, 5)
        
        sizer.Add(btn_sizer, 0, wx.EXPAND | wx.ALL, 10)
        
        panel.SetSizer(sizer)


###########################################################################
## 즐겨찾기 다이얼로그
###########################################################################
class FavoritesDialog(wx.Dialog):
    def __init__(self, parent):
        super().__init__(parent, title="즐겨찾기", size=(600, 500))
        self.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        
        self.favorites = [
            ('지출', '식비', '15000', '점심'),
            ('지출', '교통비', '5000', '버스'),
            ('수입', '급여', '3000000', '월급'),
        ]
        
        panel = wx.Panel(self)
        panel.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 리스트
        list_card = CardPanel(panel, "즐겨찾기 목록")
        self.list = wx.ListCtrl(list_card, style=wx.LC_REPORT | wx.BORDER_SIMPLE)
        self.list.SetBackgroundColour(wx.WHITE)
        
        self.list.InsertColumn(0, "구분", width=80)
        self.list.InsertColumn(1, "항목", width=150)
        self.list.InsertColumn(2, "금액", width=120)
        self.list.InsertColumn(3, "비고", width=200)
        
        for fav in self.favorites:
            idx = self.list.InsertItem(self.list.GetItemCount(), fav[0])
            self.list.SetItem(idx, 1, fav[1])
            self.list.SetItem(idx, 2, fav[2])
            self.list.SetItem(idx, 3, fav[3])
        
        list_card.AddContent(self.list)
        sizer.Add(list_card, 1, wx.EXPAND | wx.ALL, 10)
        
        # 버튼
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        btn_sizer.AddStretchSpacer()
        
        btn_cancel = wx.Button(panel, wx.ID_CANCEL, "취소", size=(100, 36))
        btn_apply = wx.Button(panel, wx.ID_OK, "적용", size=(100, 36))
        btn_apply.SetBackgroundColour(ColorTheme.ACCENT)
        btn_apply.SetForegroundColour(wx.WHITE)
        
        btn_sizer.Add(btn_cancel, 0, wx.ALL, 5)
        btn_sizer.Add(btn_apply, 0, wx.ALL, 5)
        
        sizer.Add(btn_sizer, 0, wx.EXPAND | wx.ALL, 10)
        
        panel.SetSizer(sizer)
    
    def GetSelectedFavorite(self):
        idx = self.list.GetFirstSelected()
        if idx == -1:
            return None
        
        return (
            self.list.GetItemText(idx, 0),
            self.list.GetItemText(idx, 1),
            self.list.GetItemText(idx, 2),
            self.list.GetItemText(idx, 3)
        )


###########################################################################
## 메인 프레임
###########################################################################
class MyFrame(wx.Frame):
    def __init__(self, parent):
        super().__init__(parent, title="스마트 가계부 v6.0", size=(1200, 800))
        self.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        
        # 메뉴바
        self.InitMenuBar()
        
        # 메인 패널
        main_panel = wx.Panel(self)
        main_panel.SetBackgroundColour(ColorTheme.BG_PRIMARY)
        main_sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 상단 정보 패널
        info_panel = self.CreateInfoPanel(main_panel)
        main_sizer.Add(info_panel, 0, wx.EXPAND | wx.ALL, 10)
        
        # 중간 컨텐츠
        content_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        # 좌측: 입력 패널
        left_panel = self.CreateInputPanel(main_panel)
        content_sizer.Add(left_panel, 1, wx.EXPAND | wx.ALL, 5)
        
        # 우측: 그래프 패널
        right_panel = self.CreateGraphPanel(main_panel)
        content_sizer.Add(right_panel, 1, wx.EXPAND | wx.ALL, 5)
        
        main_sizer.Add(content_sizer, 1, wx.EXPAND)
        
        # 하단: 리스트 패널
        list_panel = self.CreateListPanel(main_panel)
        main_sizer.Add(list_panel, 2, wx.EXPAND | wx.ALL, 10)
        
        main_panel.SetSizer(main_sizer)
        
        self.Centre()
        
        # 초기 데이터 로드
        self.OnSelectAll(None)
    
    def InitMenuBar(self):
        menubar = wx.MenuBar()
        
        # 파일 메뉴
        file_menu = wx.Menu()
        file_menu.Append(wx.ID_ANY, "내보내기\tCtrl+E")
        file_menu.Append(wx.ID_ANY, "가져오기\tCtrl+I")
        file_menu.AppendSeparator()
        file_menu.Append(wx.ID_EXIT, "종료\tCtrl+Q")
        menubar.Append(file_menu, "파일")
        
        # 편집 메뉴
        edit_menu = wx.Menu()
        edit_menu.Append(wx.ID_ANY, "검색\tCtrl+F")
        edit_menu.Append(wx.ID_ANY, "통계\tCtrl+S")
        menubar.Append(edit_menu, "편집")
        
        # 도구 메뉴
        tools_menu = wx.Menu()
        tools_menu.Append(wx.ID_ANY, "예산 관리")
        tools_menu.Append(wx.ID_ANY, "즐겨찾기")
        menubar.Append(tools_menu, "도구")
        
        self.SetMenuBar(menubar)
        
        # 이벤트 바인딩
        self.Bind(wx.EVT_MENU, self.OnExport, id=file_menu.FindItem("내보내기\tCtrl+E"))
        self.Bind(wx.EVT_MENU, self.OnImport, id=file_menu.FindItem("가져오기\tCtrl+I"))
        self.Bind(wx.EVT_MENU, self.OnExit, id=wx.ID_EXIT)
        self.Bind(wx.EVT_MENU, self.OnSearch, id=edit_menu.FindItem("검색\tCtrl+F"))
        self.Bind(wx.EVT_MENU, self.OnStatistics, id=edit_menu.FindItem("통계\tCtrl+S"))
        self.Bind(wx.EVT_MENU, self.OnBudget, id=tools_menu.FindItem("예산 관리"))
        self.Bind(wx.EVT_MENU, self.OnFavorites, id=tools_menu.FindItem("즐겨찾기"))
    
    def CreateInfoPanel(self, parent):
        panel = CardPanel(parent, "월간 요약")
        
        content = wx.Panel(panel)
        content.SetBackgroundColour(ColorTheme.CARD_BG)
        sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        # 월 선택
        month_box = wx.BoxSizer(wx.VERTICAL)
        month_label = wx.StaticText(content, label="조회 월")
        month_label.SetForegroundColour(ColorTheme.TEXT_SECONDARY)
        month_box.Add(month_label, 0, wx.BOTTOM, 5)
        
        self.comboMonth = wx.ComboBox(content, choices=HL_CRUD.selectMonthList(), 
                                      style=wx.CB_READONLY, size=(150, -1))
        if self.comboMonth.GetCount() > 0:
            self.comboMonth.SetSelection(0)
        month_box.Add(self.comboMonth, 0, wx.EXPAND)
        
        sizer.Add(month_box, 0, wx.ALL, 10)
        sizer.AddSpacer(20)
        
        # 수입 요약
        income_box = wx.BoxSizer(wx.VERTICAL)
        income_label = wx.StaticText(content, label="총 수입")
        income_label.SetForegroundColour(ColorTheme.TEXT_SECONDARY)
        income_box.Add(income_label, 0, wx.BOTTOM, 5)
        
        self.lblIncome = wx.StaticText(content, label="0원")
        font = self.lblIncome.GetFont()
        font.SetPointSize(18)
        font.SetWeight(wx.FONTWEIGHT_BOLD)
        self.lblIncome.SetFont(font)
        self.lblIncome.SetForegroundColour(ColorTheme.INCOME_COLOR)
        income_box.Add(self.lblIncome)
        
        sizer.Add(income_box, 0, wx.ALL, 10)
        sizer.AddSpacer(20)
        
        # 지출 요약
        expense_box = wx.BoxSizer(wx.VERTICAL)
        expense_label = wx.StaticText(content, label="총 지출")
        expense_label.SetForegroundColour(ColorTheme.TEXT_SECONDARY)
        expense_box.Add(expense_label, 0, wx.BOTTOM, 5)
        
        self.lblExpense = wx.StaticText(content, label="0원")
        font = self.lblExpense.GetFont()
        font.SetPointSize(18)
        font.SetWeight(wx.FONTWEIGHT_BOLD)
        self.lblExpense.SetFont(font)
        self.lblExpense.SetForegroundColour(ColorTheme.EXPENSE_COLOR)
        expense_box.Add(self.lblExpense)
        
        sizer.Add(expense_box, 0, wx.ALL, 10)
        sizer.AddSpacer(20)
        
        # 수지 요약
        balance_box = wx.BoxSizer(wx.VERTICAL)
        balance_label = wx.StaticText(content, label="수지")
        balance_label.SetForegroundColour(ColorTheme.TEXT_SECONDARY)
        balance_box.Add(balance_label, 0, wx.BOTTOM, 5)
        
        self.lblBalance = wx.StaticText(content, label="0원")
        font = self.lblBalance.GetFont()
        font.SetPointSize(18)
        font.SetWeight(wx.FONTWEIGHT_BOLD)
        self.lblBalance.SetFont(font)
        balance_box.Add(self.lblBalance)
        
        sizer.Add(balance_box, 0, wx.ALL, 10)
        
        content.SetSizer(sizer)
        panel.AddContent(content)
        
        # 이벤트
        self.comboMonth.Bind(wx.EVT_COMBOBOX, self.OnMonthChanged)
        
        return panel
    
    def CreateInputPanel(self, parent):
        panel = CardPanel(parent, "거래 입력")
        
        content = wx.Panel(panel)
        content.SetBackgroundColour(ColorTheme.CARD_BG)
        sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 날짜
        date_sizer = wx.BoxSizer(wx.HORIZONTAL)
        date_sizer.Add(wx.StaticText(content, label="날짜:", size=(80, -1)), 0, wx.ALIGN_CENTER_VERTICAL)
        self.dateCtrl = wx.adv.DatePickerCtrl(content, style=wx.adv.DP_DROPDOWN | wx.adv.DP_SHOWCENTURY)
        self.dateCtrl.SetValue(wx.DateTime.Today())
        date_sizer.Add(self.dateCtrl, 1, wx.EXPAND)
        sizer.Add(date_sizer, 0, wx.EXPAND | wx.ALL, 5)
        
        # 노트북 (수입/지출 탭)
        self.notebook = wx.Notebook(content)
        
        # 수입 탭
        revenue_panel = wx.Panel(self.notebook)
        revenue_panel.SetBackgroundColour(wx.WHITE)
        revenue_sizer = wx.BoxSizer(wx.VERTICAL)
        
        rev_item_sizer = wx.BoxSizer(wx.HORIZONTAL)
        rev_item_sizer.Add(wx.StaticText(revenue_panel, label="항목:", size=(80, -1)), 0, wx.ALIGN_CENTER_VERTICAL)
        self.comboRevenue = wx.ComboBox(revenue_panel, 
                                        choices=['급여', '보너스', '용돈', '기타수입'],
                                        style=wx.CB_DROPDOWN)
        rev_item_sizer.Add(self.comboRevenue, 1, wx.EXPAND)
        revenue_sizer.Add(rev_item_sizer, 0, wx.EXPAND | wx.ALL, 5)
        
        rev_amount_sizer = wx.BoxSizer(wx.HORIZONTAL)
        rev_amount_sizer.Add(wx.StaticText(revenue_panel, label="금액:", size=(80, -1)), 0, wx.ALIGN_CENTER_VERTICAL)
        self.txtRevenue = wx.TextCtrl(revenue_panel)
        rev_amount_sizer.Add(self.txtRevenue, 1, wx.EXPAND)
        revenue_sizer.Add(rev_amount_sizer, 0, wx.EXPAND | wx.ALL, 5)
        
        revenue_panel.SetSizer(revenue_sizer)
        self.notebook.AddPage(revenue_panel, "수입")
        
        # 지출 탭
        expense_panel = wx.Panel(self.notebook)
        expense_panel.SetBackgroundColour(wx.WHITE)
        expense_sizer = wx.BoxSizer(wx.VERTICAL)
        
        exp_item_sizer = wx.BoxSizer(wx.HORIZONTAL)
        exp_item_sizer.Add(wx.StaticText(expense_panel, label="항목:", size=(80, -1)), 0, wx.ALIGN_CENTER_VERTICAL)
        self.comboExpense = wx.ComboBox(expense_panel,
                                        choices=['식비', '교통비', '통신비', '주거비', '의류', '문화생활', '기타지출'],
                                        style=wx.CB_DROPDOWN)
        exp_item_sizer.Add(self.comboExpense, 1, wx.EXPAND)
        expense_sizer.Add(exp_item_sizer, 0, wx.EXPAND | wx.ALL, 5)
        
        exp_amount_sizer = wx.BoxSizer(wx.HORIZONTAL)
        exp_amount_sizer.Add(wx.StaticText(expense_panel, label="금액:", size=(80, -1)), 0, wx.ALIGN_CENTER_VERTICAL)
        self.txtExpense = wx.TextCtrl(expense_panel)
        exp_amount_sizer.Add(self.txtExpense, 1, wx.EXPAND)
        expense_sizer.Add(exp_amount_sizer, 0, wx.EXPAND | wx.ALL, 5)
        
        expense_panel.SetSizer(expense_sizer)
        self.notebook.AddPage(expense_panel, "지출")
        
        sizer.Add(self.notebook, 0, wx.EXPAND | wx.ALL, 5)
        
        # 비고
        remark_sizer = wx.BoxSizer(wx.HORIZONTAL)
        remark_sizer.Add(wx.StaticText(content, label="비고:", size=(80, -1)), 0, wx.ALIGN_CENTER_VERTICAL)
        self.txtRemark = wx.TextCtrl(content, style=wx.TE_MULTILINE, size=(-1, 60))
        remark_sizer.Add(self.txtRemark, 1, wx.EXPAND)
        sizer.Add(remark_sizer, 1, wx.EXPAND | wx.ALL, 5)
        
        # 버튼들
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        self.btnInsert = wx.Button(content, label="추가", size=(100, 36))
        self.btnInsert.SetBackgroundColour(ColorTheme.ACCENT)
        self.btnInsert.SetForegroundColour(wx.WHITE)
        btn_sizer.Add(self.btnInsert, 0, wx.ALL, 5)
        
        self.btnUpdate = wx.Button(content, label="수정", size=(100, 36))
        btn_sizer.Add(self.btnUpdate, 0, wx.ALL, 5)
        
        self.btnDelete = wx.Button(content, label="삭제", size=(100, 36))
        btn_sizer.Add(self.btnDelete, 0, wx.ALL, 5)
        
        self.btnClear = wx.Button(content, label="초기화", size=(100, 36))
        btn_sizer.Add(self.btnClear, 0, wx.ALL, 5)
        
        sizer.Add(btn_sizer, 0, wx.ALIGN_CENTER | wx.ALL, 10)
        
        content.SetSizer(sizer)
        panel.AddContent(content)
        
        # 이벤트 바인딩
        self.btnInsert.Bind(wx.EVT_BUTTON, self.OnInsert)
        self.btnUpdate.Bind(wx.EVT_BUTTON, self.OnUpdate)
        self.btnDelete.Bind(wx.EVT_BUTTON, self.OnDelete)
        self.btnClear.Bind(wx.EVT_BUTTON, lambda e: self.ClearInputs())
        
        return panel
    
    def CreateGraphPanel(self, parent):
        panel = CardPanel(parent, "지출 분석")
        
        content = wx.Panel(panel)
        content.SetBackgroundColour(ColorTheme.CARD_BG)
        sizer = wx.BoxSizer(wx.VERTICAL)
        
        self.graphPanel = Barchart(content)
        sizer.Add(self.graphPanel, 1, wx.EXPAND | wx.ALL, 5)
        
        btn_graph = wx.Button(content, label="그래프 생성", size=(150, 36))
        btn_graph.SetBackgroundColour(ColorTheme.ACCENT)
        btn_graph.SetForegroundColour(wx.WHITE)
        sizer.Add(btn_graph, 0, wx.ALIGN_CENTER | wx.ALL, 10)
        
        content.SetSizer(sizer)
        panel.AddContent(content)
        
        btn_graph.Bind(wx.EVT_BUTTON, self.OnMakeGraph)
        
        return panel
    
    def CreateListPanel(self, parent):
        panel = CardPanel(parent, "거래 내역")
        
        content = wx.Panel(panel)
        content.SetBackgroundColour(ColorTheme.CARD_BG)
        sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 버튼들
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        btn_all = wx.Button(content, label="전체 조회", size=(100, 32))
        btn_sizer.Add(btn_all, 0, wx.ALL, 5)
        
        btn_month = wx.Button(content, label="월별 조회", size=(100, 32))
        btn_sizer.Add(btn_month, 0, wx.ALL, 5)
        
        btn_load = wx.Button(content, label="불러오기", size=(100, 32))
        btn_sizer.Add(btn_load, 0, wx.ALL, 5)
        
        btn_sizer.AddStretchSpacer()
        sizer.Add(btn_sizer, 0, wx.EXPAND | wx.BOTTOM, 10)
        
        # 리스트
        self.list = wx.ListCtrl(content, style=wx.LC_REPORT | wx.BORDER_SIMPLE)
        self.list.SetBackgroundColour(wx.WHITE)
        
        self.list.InsertColumn(0, "번호", width=80)
        self.list.InsertColumn(1, "날짜", width=100)
        self.list.InsertColumn(2, "구분", width=80)
        self.list.InsertColumn(3, "항목", width=150)
        self.list.InsertColumn(4, "수입", width=120)
        self.list.InsertColumn(5, "지출", width=120)
        self.list.InsertColumn(6, "비고", width=300)
        
        sizer.Add(self.list, 1, wx.EXPAND)
        
        content.SetSizer(sizer)
        panel.AddContent(content)
        
        # 이벤트
        btn_all.Bind(wx.EVT_BUTTON, self.OnSelectAll)
        btn_month.Bind(wx.EVT_BUTTON, self.OnSelectMonth)
        btn_load.Bind(wx.EVT_BUTTON, self.OnLoadToInput)
        
        return panel
    
    def OnMonthChanged(self, event):
        self.OnSelectMonth(None)
    
    def OnInsert(self, event):
        date_str = self.dateCtrl.GetValue().FormatISODate()
        
        if self.notebook.GetSelection() == 0:  # 수입
            section = '수입'
            title = self.comboRevenue.GetValue()
            revenue = self.txtRevenue.GetValue().replace(',', '')
            expense = '0'
        else:  # 지출
            section = '지출'
            title = self.comboExpense.GetValue()
            revenue = '0'
            expense = self.txtExpense.GetValue().replace(',', '')
        
        remark = self.txtRemark.GetValue()
        
        if not title:
            wx.MessageBox("항목을 입력하세요.", "입력 오류", wx.OK | wx.ICON_WARNING)
            return
        
        if revenue == '0' and expense == '0':
            wx.MessageBox("금액을 입력하세요.", "입력 오류", wx.OK | wx.ICON_WARNING)
            return
        
        try:
            HL_CRUD.insert((date_str, section, title, revenue, expense, remark))
            wx.MessageBox("추가되었습니다.", "성공", wx.OK | wx.ICON_INFORMATION)
            self.OnSelectAll(None)
            self.ClearInputs()
        except Exception as e:
            wx.MessageBox(f"추가 실패: {str(e)}", "오류", wx.OK | wx.ICON_ERROR)
    
    def OnUpdate(self, event):
        idx = self.list.GetFirstSelected()
        if idx == -1:
            wx.MessageBox("수정할 항목을 선택하세요.", "알림", wx.OK | wx.ICON_WARNING)
            return
        
        key = self.list.GetItemText(idx, 0)
        date_str = self.dateCtrl.GetValue().FormatISODate()
        
        if self.notebook.GetSelection() == 0:
            section = '수입'
            title = self.comboRevenue.GetValue()
            revenue = self.txtRevenue.GetValue().replace(',', '')
            expense = '0'
        else:
            section = '지출'
            title = self.comboExpense.GetValue()
            revenue = '0'
            expense = self.txtExpense.GetValue().replace(',', '')
        
        remark = self.txtRemark.GetValue()
        
        try:
            HL_CRUD.update((key, date_str, section, title, revenue, expense, remark))
            wx.MessageBox("수정되었습니다.", "성공", wx.OK | wx.ICON_INFORMATION)
            self.OnSelectAll(None)
            self.ClearInputs()
        except Exception as e:
            wx.MessageBox(f"수정 실패: {str(e)}", "오류", wx.OK | wx.ICON_ERROR)
    
    def OnDelete(self, event):
        idx = self.list.GetFirstSelected()
        if idx == -1:
            wx.MessageBox("삭제할 항목을 선택하세요.", "알림", wx.OK | wx.ICON_WARNING)
            return
        
        if wx.MessageBox("정말 삭제하시겠습니까?", "삭제 확인", wx.YES_NO | wx.ICON_QUESTION) == wx.YES:
            key = self.list.GetItemText(idx, 0)
            try:
                HL_CRUD.delete(key)
                wx.MessageBox("삭제되었습니다.", "성공", wx.OK | wx.ICON_INFORMATION)
                self.OnSelectAll(None)
                self.ClearInputs()
            except Exception as e:
                wx.MessageBox(f"삭제 실패: {str(e)}", "오류", wx.OK | wx.ICON_ERROR)
    
    def OnSelectAll(self, event):
        self.list.DeleteAllItems()
        rows = HL_CRUD.selectAll()
        
        for row in rows:
            idx = self.list.InsertItem(self.list.GetItemCount(), str(row[0]))
            self.list.SetItem(idx, 1, row[1])
            self.list.SetItem(idx, 2, row[2])
            self.list.SetItem(idx, 3, row[3])
            self.list.SetItem(idx, 4, f"{float(row[4]):,.0f}" if row[4] else "0")
            self.list.SetItem(idx, 5, f"{float(row[5]):,.0f}" if row[5] else "0")
            self.list.SetItem(idx, 6, row[6])
            
            # 색상
            if row[2] == '수입':
                self.list.SetItemTextColour(idx, ColorTheme.INCOME_COLOR)
            else:
                self.list.SetItemTextColour(idx, ColorTheme.EXPENSE_COLOR)
    
    def OnSelectMonth(self, event):
        month = self.comboMonth.GetValue()
        if not month:
            wx.MessageBox("월을 선택하세요.", "알림", wx.OK | wx.ICON_WARNING)
            return
        
        self.list.DeleteAllItems()
        rows = HL_CRUD.selectMonthlySum(month)
        
        total_income = 0
        total_expense = 0
        
        for row in rows:
            if row[2] == '합계':
                total_income = float(row[4]) if row[4] else 0
                total_expense = float(row[5]) if row[5] else 0
                continue
            
            idx = self.list.InsertItem(self.list.GetItemCount(), str(row[0]))
            self.list.SetItem(idx, 1, row[1])
            self.list.SetItem(idx, 2, row[2])
            self.list.SetItem(idx, 3, row[3])
            self.list.SetItem(idx, 4, f"{float(row[4]):,.0f}" if row[4] else "0")
            self.list.SetItem(idx, 5, f"{float(row[5]):,.0f}" if row[5] else "0")
            self.list.SetItem(idx, 6, row[6])
            
            if row[2] == '수입':
                self.list.SetItemTextColour(idx, ColorTheme.INCOME_COLOR)
            else:
                self.list.SetItemTextColour(idx, ColorTheme.EXPENSE_COLOR)
        
        # 요약 업데이트
        self.lblIncome.SetLabel(f"{total_income:,.0f}원")
        self.lblExpense.SetLabel(f"{total_expense:,.0f}원")
        
        balance = total_income - total_expense
        self.lblBalance.SetLabel(f"{balance:,.0f}원")
        
        if balance >= 0:
            self.lblBalance.SetForegroundColour(ColorTheme.INCOME_COLOR)
        else:
            self.lblBalance.SetForegroundColour(ColorTheme.EXPENSE_COLOR)
    
    def OnLoadToInput(self, event):
        idx = self.list.GetFirstSelected()
        if idx == -1:
            wx.MessageBox("불러올 항목을 선택하세요.", "알림", wx.OK | wx.ICON_WARNING)
            return
        
        date_str = self.list.GetItemText(idx, 1)
        section = self.list.GetItemText(idx, 2)
        title = self.list.GetItemText(idx, 3)
        revenue = self.list.GetItemText(idx, 4).replace(',', '')
        expense = self.list.GetItemText(idx, 5).replace(',', '')
        remark = self.list.GetItemText(idx, 6)
        
        # 날짜 설정
        try:
            date = wx.DateTime()
            date.ParseDate(date_str)
            self.dateCtrl.SetValue(date)
        except:
            pass
        
        # 섹션에 따라 탭 전환
        if section == '수입':
            self.notebook.SetSelection(0)
            self.comboRevenue.SetValue(title)
            self.txtRevenue.SetValue(revenue if revenue != '0' else '')
        else:
            self.notebook.SetSelection(1)
            self.comboExpense.SetValue(title)
            self.txtExpense.SetValue(expense if expense != '0' else '')
        
        self.txtRemark.SetValue(remark)
    
    def ClearInputs(self):
        self.txtRevenue.Clear()
        self.txtExpense.Clear()
        self.txtRemark.Clear()
        self.comboRevenue.SetValue('')
        self.comboExpense.SetValue('')
        self.dateCtrl.SetValue(wx.DateTime.Today())
    
    def OnMakeGraph(self, event):
        rows = HL_CRUD.selectAll()
        expense_data = defaultdict(float)
        
        for row in rows:
            if row[2] == '지출':
                title = row[3].split('.')[0] if '.' in row[3] else row[3]
                try:
                    amount = float(row[5]) if row[5] else 0
                    if amount > 0:
                        expense_data[title] += amount / 1000
                except (ValueError, TypeError):
                    continue
        
        if expense_data:
            self.graphPanel.SetData(dict(expense_data))
            wx.MessageBox("그래프가 생성되었습니다.", "그래프 생성", wx.OK | wx.ICON_INFORMATION)
        else:
            wx.MessageBox("표시할 지출 데이터가 없습니다.", "그래프 생성", wx.OK | wx.ICON_WARNING)
    
    def OnExport(self, event):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            dlg = wx.FileDialog(
                self, "Excel 파일로 저장",
                wildcard="Excel files (*.xlsx)|*.xlsx",
                style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT
            )
            
            if dlg.ShowModal() == wx.ID_OK:
                filepath = dlg.GetPath()
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "가계부"
                
                headers = ["거래번호", "날짜", "구분", "상세내역", "수입", "지출", "비고"]
                ws.append(headers)
                
                header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                rows = HL_CRUD.selectAll()
                for row in rows:
                    ws.append(list(row))
                
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 30
                
                wb.save(filepath)
                
                wx.MessageBox("Excel 파일로 저장되었습니다.", "내보내기 완료", wx.OK | wx.ICON_INFORMATION)
            
            dlg.Destroy()
            
        except ImportError:
            wx.MessageBox("openpyxl 모듈이 필요합니다.\npip install openpyxl", "모듈 오류", wx.OK | wx.ICON_ERROR)
        except Exception as e:
            wx.MessageBox(f"내보내기 실패: {str(e)}", "오류", wx.OK | wx.ICON_ERROR)
    
    def OnImport(self, event):
        dlg = wx.FileDialog(
            self, "CSV 파일 선택",
            wildcard="CSV files (*.csv)|*.csv",
            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
        )
        
        if dlg.ShowModal() == wx.ID_OK:
            filepath = dlg.GetPath()
            
            try:
                count = 0
                with open(filepath, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    next(reader)
                    
                    for row in reader:
                        if len(row) >= 6:
                            HL_CRUD.insert((row[1], row[2], row[3], row[4], row[5], row[6]))
                            count += 1
                
                wx.MessageBox(f"{count}건의 데이터를 가져왔습니다.", "가져오기 완료", wx.OK | wx.ICON_INFORMATION)
                self.OnSelectAll(None)
                
            except Exception as e:
                wx.MessageBox(f"가져오기 실패: {str(e)}", "오류", wx.OK | wx.ICON_ERROR)
        
        dlg.Destroy()
    
    def OnExit(self, event):
        self.Close()
    
    def OnSearch(self, event):
        dlg = SearchDialog(self)
        
        if dlg.ShowModal() == wx.ID_OK:
            criteria = dlg.GetSearchCriteria()
            
            self.list.DeleteAllItems()
            rows = HL_CRUD.selectAll()
            
            count = 0
            for row in rows:
                if row[1] < criteria['start_date'] or row[1] > criteria['end_date']:
                    continue
                
                if row[2] == '수입' and not criteria['include_income']:
                    continue
                if row[2] == '지출' and not criteria['include_expense']:
                    continue
                
                amount = float(row[4]) if row[4] else float(row[5]) if row[5] else 0
                
                if criteria['min_amount']:
                    try:
                        if amount < float(criteria['min_amount'].replace(',', '')):
                            continue
                    except ValueError:
                        pass
                
                if criteria['max_amount']:
                    try:
                        if amount > float(criteria['max_amount'].replace(',', '')):
                            continue
                    except ValueError:
                        pass
                
                if criteria['keyword'] and criteria['keyword'] not in row[6]:
                    continue
                
                idx = self.list.InsertItem(self.list.GetItemCount(), str(row[0]))
                self.list.SetItem(idx, 1, row[1])
                self.list.SetItem(idx, 2, row[2])
                self.list.SetItem(idx, 3, row[3])
                self.list.SetItem(idx, 4, f"{float(row[4]):,.0f}" if row[4] else "0")
                self.list.SetItem(idx, 5, f"{float(row[5]):,.0f}" if row[5] else "0")
                self.list.SetItem(idx, 6, row[6])
                count += 1
            
            wx.MessageBox(f"검색 완료 - {count}건 발견", "검색 결과", wx.OK | wx.ICON_INFORMATION)
        
        dlg.Destroy()
    
    def OnStatistics(self, event):
        rows = HL_CRUD.selectAll()
        dlg = StatisticsDialog(self, rows)
        dlg.ShowModal()
        dlg.Destroy()
    
    def OnBudget(self, event):
        dlg = BudgetDialog(self)
        dlg.ShowModal()
        dlg.Destroy()
    
    def OnFavorites(self, event):
        dlg = FavoritesDialog(self)
        
        if dlg.ShowModal() == wx.ID_OK:
            favorite = dlg.GetSelectedFavorite()
            if favorite:
                if favorite[0] == '수입':
                    self.notebook.SetSelection(0)
                    self.comboRevenue.SetValue(favorite[1])
                    self.txtRevenue.SetValue(favorite[2])
                else:
                    self.notebook.SetSelection(1)
                    self.comboExpense.SetValue(favorite[1])
                    self.txtExpense.SetValue(favorite[2])
                
                self.txtRemark.SetValue(favorite[3])
                wx.MessageBox("즐겨찾기 항목이 적용되었습니다.", "적용 완료", wx.OK | wx.ICON_INFORMATION)
        
        dlg.Destroy()


if __name__ == '__main__':
    app = wx.App()
    frame = MyFrame(parent=None)
    frame.Show()
    app.MainLoop()
